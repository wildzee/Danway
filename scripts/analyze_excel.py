#!/usr/bin/env python3
"""
Script to analyze Excel files and display their structure
"""
import sys
import json

try:
    import openpyxl
    import xlrd
except ImportError:
    print("Installing required libraries...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "xlrd"])
    import openpyxl
    import xlrd

def analyze_xlsx(file_path):
    """Analyze .xlsx file"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {file_path}")
    print(f"{'='*80}\n")
    
    wb = openpyxl.load_workbook(file_path)
    
    print(f"Number of sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}\n")
        
        # Get headers (first row)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col).value
            headers.append(cell_value)
        
        print("Headers:")
        for idx, header in enumerate(headers, 1):
            print(f"  Column {idx}: {header}")
        
        # Show first 5 rows of data
        print(f"\nFirst 5 rows of data:")
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(ws.max_column + 1, 15)):  # Limit to 15 columns for display
                cell_value = ws.cell(row_idx, col).value
                row_data.append(str(cell_value)[:30] if cell_value else "")
            print(f"  Row {row_idx}: {' | '.join(row_data)}")
        
        print(f"\nTotal rows: {ws.max_row}")

def analyze_xls(file_path):
    """Analyze .xls file"""
    print(f"\n{'='*80}")
    print(f"Analyzing: {file_path}")
    print(f"{'='*80}\n")
    
    wb = xlrd.open_workbook(file_path)
    
    print(f"Number of sheets: {wb.nsheets}")
    print(f"Sheet names: {wb.sheet_names()}\n")
    
    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Rows: {ws.nrows}, Columns: {ws.ncols}\n")
        
        # Get headers (first row)
        if ws.nrows > 0:
            headers = ws.row_values(0)
            print("Headers:")
            for idx, header in enumerate(headers, 1):
                print(f"  Column {idx}: {header}")
            
            # Show first 5 rows of data
            print(f"\nFirst 5 rows of data:")
            for row_idx in range(min(5, ws.nrows)):
                row_data = ws.row_values(row_idx)
                # Limit display to first 15 columns
                display_data = [str(val)[:30] for val in row_data[:15]]
                print(f"  Row {row_idx + 1}: {' | '.join(display_data)}")
            
            print(f"\nTotal rows: {ws.nrows}")

if __name__ == "__main__":
    # Analyze both files
    sap_file = "/Users/mdafjalkhan/DanwayEME/data/D657-SAP Attadance- 13-02-2026.xlsx"
    punch_file = "/Users/mdafjalkhan/DanwayEME/data/14-2-2026PuncReport.xls"
    
    analyze_xlsx(sap_file)
    print("\n" + "="*80 + "\n")
    analyze_xls(punch_file)
